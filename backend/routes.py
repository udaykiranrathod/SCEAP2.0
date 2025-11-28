from fastapi import APIRouter
from typing import List
from schemas import CableInput, CableOutput
from cable_engine import (
    full_load_current_kw,
    full_load_current_kva,
    derated_current,
	voltage_drop_percent,
	voltage_drop_percent_from_rx,
    short_circuit_check,
	motor_start_vdrop_percent,
	motor_start_multiplier,
)
from catalog_importer import normalize_catalog, read_excel_headers_sample, get_catalog_path, get_catalog_excel_path, load_catalog
from fastapi import UploadFile, File, HTTPException
from fastapi.responses import FileResponse
import pandas as pd
import uuid
import os
import json
from io import BytesIO

router = APIRouter(prefix="/cable", tags=["Cable Sizing"])


@router.post("/size", response_model=CableOutput)
def size_cable(data: CableInput):

	# Determine base current
	if data.current > 0:
		i_base = data.current
	elif data.load_kw > 0:
		i_base = full_load_current_kw(data.load_kw, data.voltage, data.pf, data.eff)
	else:
		i_base = full_load_current_kva(data.load_kva, data.voltage)

	i_derated = derated_current(i_base, data.derating_factors)

	# Voltage drop: prefer R/X if provided (ohm/km), otherwise use mv_per_a_m (mV/A/m)
	if getattr(data, 'r_ohm_per_km', None) not in (None, 0) and getattr(data, 'x_ohm_per_km', None) not in (None, 0):
		vdrop = voltage_drop_percent_from_rx(i_base, data.length, data.r_ohm_per_km, data.x_ohm_per_km, data.voltage, data.pf or 1.0)
	else:
		vdrop = voltage_drop_percent(i_base, data.length, data.mv_per_a_m, data.voltage)

	# Select cable CSA by meeting derated current
	selected_csa = None
	for csa in sorted(data.csa_options):
		if csa >= i_derated:
			selected_csa = csa
			break

	if selected_csa is None:
		selected_csa = max(data.csa_options)

	# SC Check
	sc_ok, a_required = short_circuit_check(
		data.sc_current,
		data.sc_time,
		data.k_const,
		selected_csa,
	)

	vdrop_ok = vdrop <= 5  # default for LV

	# Compliance messages
	compliance = []
	if vdrop_ok:
		compliance.append({"type": "vdrop", "ok": True, "msg": f"Running Vdrop OK ({round(vdrop,3)}% < 5%)"})
	else:
		compliance.append({"type": "vdrop", "ok": False, "msg": f"Running Vdrop too high ({round(vdrop,3)}% > 5%)"})

	if sc_ok:
		margin = None
		try:
			margin = round(((selected_csa - a_required) / a_required) * 100, 1) if a_required > 0 else None
		except Exception:
			margin = None
		if margin is not None:
			compliance.append({"type": "sc", "ok": True, "msg": f"SC OK with {margin}% margin"})
		else:
			compliance.append({"type": "sc", "ok": True, "msg": "SC OK"})
	else:
		compliance.append({"type": "sc", "ok": False, "msg": "SC check failed (selected CSA may be insufficient)"})

	# Voltage category (simple)
	voltage_category = "LV" if data.voltage <= 1000 else "MV"

	return CableOutput(
		cable_number=data.cable_number,
		flc=round(i_base, 2),
		derated_current=round(i_derated, 2),
		selected_csa=selected_csa,
		vdrop_percent=round(vdrop, 3),
		sc_required_area=round(a_required, 2),
		sc_ok=sc_ok,
		vdrop_ok=vdrop_ok,
		compliance=compliance,
		voltage_category=voltage_category,
	)


@router.post("/bulk-size", response_model=List[CableOutput])
def bulk_size_cables(data: List[CableInput]):
	results: List[CableOutput] = []

	for cable in data:
		# Determine base current
		if cable.current and cable.current > 0:
			i_base = cable.current
		elif cable.load_kw and cable.load_kw > 0:
			i_base = full_load_current_kw(cable.load_kw, cable.voltage, cable.pf or 1.0, cable.eff or 1.0)
		else:
			i_base = full_load_current_kva(cable.load_kva or 0, cable.voltage)

		i_derated = derated_current(i_base, cable.derating_factors or [1.0])

		# Voltage drop: prefer R/X if present
		if getattr(cable, 'r_ohm_per_km', None) not in (None, 0) and getattr(cable, 'x_ohm_per_km', None) not in (None, 0):
			vdrop = voltage_drop_percent_from_rx(i_base, cable.length, cable.r_ohm_per_km, cable.x_ohm_per_km, cable.voltage, cable.pf or 1.0)
		else:
			vdrop = voltage_drop_percent(i_base, cable.length, cable.mv_per_a_m, cable.voltage)

		# Select cable CSA by meeting derated current
		selected_csa = None
		for csa in sorted(cable.csa_options or []):
			if csa >= i_derated:
				selected_csa = csa
				break

		if selected_csa is None and cable.csa_options:
			selected_csa = max(cable.csa_options)

		# SC Check
		sc_ok, a_required = short_circuit_check(
			cable.sc_current or 0,
			cable.sc_time or 1,
			cable.k_const or 115,
			selected_csa or 0,
		)

		# Determine vdrop limits (basic defaults)
		vdrop_limit = 5 if ((cable.voltage or 0) <= 1000) else 10
		vdrop_ok = vdrop <= vdrop_limit

		# Motor starting vdrop (uses motor_start_method from payload, default DOL)
		start_method = getattr(cable, 'motor_start_method', 'DOL') or 'DOL'
		try:
			start_vdrop, start_current = motor_start_vdrop_percent(
				i_base,
				start_method,
				cable.length or 0,
				cable.r_ohm_per_km or 0,
				cable.x_ohm_per_km or 0,
				cable.voltage or 0,
				cable.pf or 1.0,
			)
		except Exception:
			start_vdrop, start_current = 0.0, 0.0
		start_limit = 15.0  # default start vdrop allowable percent for motors
		start_ok = start_vdrop <= start_limit

		# Build compliance messages with richer payload: type, ok, limit, value, margin, msg
		compliance = []

		# Running voltage drop
		vdrop_margin = None
		try:
			vdrop_margin = round(((vdrop_limit - vdrop) / vdrop_limit) * 100, 1) if vdrop_limit and vdrop is not None else None
		except Exception:
			vdrop_margin = None
		compliance.append({
			"type": "vdrop_run",
			"ok": bool(vdrop_ok),
			"limit": round(vdrop_limit, 2),
			"value": round(vdrop, 3),
			"margin": vdrop_margin,
			"msg": f"Running Vdrop {round(vdrop,3)}% {'OK' if vdrop_ok else 'EXCEEDS'} limit {vdrop_limit}%"
		})

		# Starting voltage drop (motor)
		start_margin = None
		try:
			start_margin = round(((start_limit - start_vdrop) / start_limit) * 100, 1) if start_limit and start_vdrop is not None else None
		except Exception:
			start_margin = None
		compliance.append({
			"type": "vdrop_start",
			"ok": bool(start_ok),
			"limit": round(start_limit, 2),
			"value": round(start_vdrop, 3),
			"margin": start_margin,
			"msg": f"Start Vdrop {round(start_vdrop,3)}% via {start_method} {'OK' if start_ok else 'EXCEEDS'} {start_limit}%"
		})

		# Short circuit withstand
		sc_margin = None
		try:
			if a_required > 0:
				sc_margin = round(((selected_csa or 0) - a_required) / a_required * 100, 1)
		except Exception:
			sc_margin = None
		compliance.append({
			"type": "sc",
			"ok": bool(sc_ok),
			"limit": round(a_required, 2) if a_required is not None else None,
			"value": round(selected_csa or 0, 2),
			"margin": sc_margin,
			"msg": "SC OK" if sc_ok else "SC check failed (selected CSA may be insufficient)"
		})

		# Derating check (simple: ensure derated_current <= selected_csa rating placeholder)
		# Note: selected_csa may represent an available conductor rating for this project.
		derate_ok = True
		derate_margin = None
		try:
			if selected_csa and selected_csa > 0:
				derate_ok = (i_derated <= selected_csa)
				derate_margin = round(((selected_csa - i_derated) / (selected_csa or 1)) * 100, 1)
		except Exception:
			derate_ok = False

		compliance.append({
			"type": "derating",
			"ok": bool(derate_ok),
			"limit": round(selected_csa or 0, 2),
			"value": round(i_derated, 2),
			"margin": derate_margin,
			"msg": "Derating OK" if derate_ok else "Derating exceeds selected conductor capacity"
		})


		# Grouping-factor check: product of derating factors indicates grouping/installation derating
		grouping_factor = 1.0
		try:
			for f in (cable.derating_factors or [1.0]):
				grouping_factor *= float(f or 1.0)
		except Exception:
			grouping_factor = 1.0

		# Use optional per-request grouping_threshold when provided, otherwise default
		grouping_threshold = float(getattr(cable, 'grouping_threshold', 0.85) or 0.85)
		grouping_ok = grouping_factor >= grouping_threshold
		grouping_margin = None
		try:
			grouping_margin = round((1 - grouping_factor) * 100, 1)
		except Exception:
			grouping_margin = None

		compliance.append({
			"type": "grouping",
			"ok": bool(grouping_ok),
			"limit": round(grouping_threshold, 3),
			"value": round(grouping_factor, 3),
			"margin": grouping_margin,
			"msg": f"Grouping factor {round(grouping_factor,3)} ({grouping_margin}% reduction)"
		})


		# Thermal rating check: prefer catalog rated currents if supplied by frontend, otherwise fall back to selected_csa proxy
		rated_current = None
		# prefer air rating, then trench, then duct, then selected_csa as a very rough proxy
		try:
			rated_current = float(getattr(cable, 'catalog_rated_current_air', None) or getattr(cable, 'catalog_rated_current_trench', None) or getattr(cable, 'catalog_rated_current_duct', None) or 0)
		except Exception:
			rated_current = 0

		if not rated_current and selected_csa:
			# simple proxy: map CSA to a nominal current (existing behavior); keep using selected_csa as proxy
			rated_current = float(selected_csa or 0)

		thermal_ok = True
		thermal_margin = None
		try:
			if rated_current and rated_current > 0:
				thermal_ok = (i_derated <= rated_current)
				thermal_margin = round(((rated_current - i_derated) / (rated_current or 1)) * 100, 1)
		except Exception:
			thermal_ok = False

		compliance.append({
			"type": "thermal",
			"ok": bool(thermal_ok),
			"limit": round(rated_current or 0, 2),
			"value": round(i_derated, 2),
			"margin": thermal_margin,
			"msg": "Thermal rating OK" if thermal_ok else "Thermal rating exceeded for selected conductor"
		})

		voltage_category = "LV" if (cable.voltage or 0) <= 1000 else "MV"

		results.append(
			CableOutput(
				cable_number=cable.cable_number or "",
				flc=round(i_base, 2),
				derated_current=round(i_derated, 2),
				selected_csa=selected_csa or 0,
				vdrop_percent=round(vdrop, 3),
				sc_required_area=round(a_required, 2),
				sc_ok=bool(sc_ok),
				vdrop_ok=bool(vdrop_ok),
				compliance=compliance,
				voltage_category=voltage_category,
			)
		)

	return results


@router.post('/export-excel', response_model=dict)
def export_excel(payload: dict):
	"""
	Accepts { rows: List[BulkRow with results], timestamp: str } 
	and returns { downloadUrl: str, filename: str } for frontend to download.
	Generates XLSX with Cable Schedule and BOQ Summary sheets.
	"""
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
		from datetime import datetime
	except ImportError:
		raise HTTPException(status_code=500, detail='openpyxl not installed')

	rows = payload.get('rows', [])
	if not rows:
		raise HTTPException(status_code=400, detail='No rows provided')

	# Create workbook
	wb = Workbook()
	
	# Remove default sheet and create new ones
	if 'Sheet' in wb.sheetnames:
		wb.remove(wb['Sheet'])

	# ===== Cable Schedule Sheet =====
	ws_schedule = wb.create_sheet('Cable Schedule', 0)
	
	headers = [
		'Cable No', 'From', 'To', 'Voltage (V)', 'Load (kW)', 'Length (m)',
		'FLC (A)', 'Derated (A)', 'CSA (mm²)', 'Vdrop (%)', 'SC OK', 'Status'
	]
	ws_schedule.append(headers)
	
	# Style header
	header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
	header_font = Font(bold=True, color='FFFFFF')
	for cell in ws_schedule[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center', vertical='center')
	
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	
	# Add data rows
	for row_data in rows:
		cable_no = row_data.get('cable_number', '')
		from_eq = row_data.get('from_equipment', '')
		to_eq = row_data.get('to_equipment', '')
		voltage = row_data.get('voltage', 415)
		load_kw = row_data.get('load_kw', 0)
		length = row_data.get('length', 0)
		
		result = row_data.get('result', {})
		flc = result.get('flc', 0)
		derated = result.get('derated_current', 0)
		csa = result.get('selected_csa', 0)
		vdrop = result.get('vdrop_percent', 0)
		sc_ok = result.get('sc_ok', False)
		vdrop_ok = result.get('vdrop_ok', False)
		
		status = 'PASS' if (sc_ok and vdrop_ok) else 'FAIL'
		status_color = 'C6EFCE' if status == 'PASS' else 'FFC7CE'
		
		ws_schedule.append([
			cable_no, from_eq, to_eq, voltage, load_kw, length,
			round(flc, 2), round(derated, 2), csa, round(vdrop, 3), 'Yes' if sc_ok else 'No', status
		])
		
		# Apply border and status color to last row
		for idx, cell in enumerate(ws_schedule[ws_schedule.max_row]):
			cell.border = thin_border
			cell.alignment = Alignment(horizontal='center', vertical='center')
			if idx == len(headers) - 1:  # Status column
				cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
	
	# Set column widths
	ws_schedule.column_dimensions['A'].width = 12
	ws_schedule.column_dimensions['B'].width = 15
	ws_schedule.column_dimensions['C'].width = 15
	ws_schedule.column_dimensions['D'].width = 12
	ws_schedule.column_dimensions['E'].width = 12
	ws_schedule.column_dimensions['F'].width = 12
	ws_schedule.column_dimensions['G'].width = 12
	ws_schedule.column_dimensions['H'].width = 12
	ws_schedule.column_dimensions['I'].width = 12
	ws_schedule.column_dimensions['J'].width = 12
	ws_schedule.column_dimensions['K'].width = 10
	ws_schedule.column_dimensions['L'].width = 10
	
	# ===== BOQ Summary Sheet =====
	ws_boq = wb.create_sheet('BOQ Summary', 1)
	
	boq_headers = ['CSA (mm²)', 'Quantity', 'Total Length (m)']
	ws_boq.append(boq_headers)
	
	for cell in ws_boq[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center', vertical='center')
	
	# Aggregate by CSA
	from collections import defaultdict
	boq_data = defaultdict(lambda: {'count': 0, 'length': 0})
	
	for row_data in rows:
		result = row_data.get('result', {})
		csa = result.get('selected_csa', 0)
		length = row_data.get('length', 0)
		
		if csa > 0:
			boq_data[csa]['count'] += 1
			boq_data[csa]['length'] += length
	
	# Add BOQ rows
	for csa in sorted(boq_data.keys()):
		count = boq_data[csa]['count']
		total_length = boq_data[csa]['length']
		ws_boq.append([csa, count, round(total_length, 2)])
		
		for cell in ws_boq[ws_boq.max_row]:
			cell.border = thin_border
			cell.alignment = Alignment(horizontal='center', vertical='center')
	
	ws_boq.column_dimensions['A'].width = 15
	ws_boq.column_dimensions['B'].width = 12
	ws_boq.column_dimensions['C'].width = 18
	
	# Save to bytes
	excel_bytes = BytesIO()
	wb.save(excel_bytes)
	excel_bytes.seek(0)
	
	# Save to temp file for download
	timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
	filename = f"SCEAP_Cable_Schedule_{timestamp}.xlsx"
	filepath = f"/tmp/{filename}"
	
	with open(filepath, 'wb') as f:
		f.write(excel_bytes.getvalue())
	
	return {
		'downloadUrl': f'/cable/download-excel/{filename}',
		'filename': filename
	}


@router.post('/export/boq', response_model=dict)
def export_boq(payload: dict):
	"""Generate BOQ XLSX grouped by vendor, part_no, cores, csa.
	Payload: { rows: [ BulkRow-like objects ], project: optional string }
	Returns: { downloadUrl, filename }
	"""
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
		from datetime import datetime
	except ImportError:
		raise HTTPException(status_code=500, detail='openpyxl not installed')

	rows = payload.get('rows', [])
	project = payload.get('project', 'PROJECT')
	if not rows:
		raise HTTPException(status_code=400, detail='No rows provided')

	wb = Workbook()
	if 'Sheet' in wb.sheetnames:
		wb.remove(wb['Sheet'])

	ws = wb.create_sheet('BOQ', 0)
	headers = ['Vendor', 'Part No', 'Spec', 'Cores', 'CSA (mm²)', 'Qty (m)', 'Weight (kg)', 'Remarks']
	ws.append(headers)

	header_fill = PatternFill(start_color='0B5394', end_color='0B5394', fill_type='solid')
	header_font = Font(bold=True, color='FFFFFF')
	thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center', vertical='center')

	from collections import defaultdict
	# key: (vendor, part_no, spec, cores, csa)
	agg = defaultdict(lambda: {'length': 0.0, 'weight': 0.0, 'remarks': set(), 'qty_count': 0})

	for r in rows:
		vendor = (r.get('catalog_vendor') or r.get('vendor') or 'UNKNOWN')
		part = (r.get('catalog_part_no') or r.get('part_no') or '—')
		spec = r.get('catalog_part_no') or ''
		cores = r.get('catalog_cores') or r.get('cores') or (r.get('result') or {}).get('cores') or ''
		csa = (r.get('catalog_csa_mm2') or (r.get('result') or {}).get('selected_csa') or 0)
		length = float(r.get('length') or 0)
		weight_per_km = float(r.get('catalog_weight_kg_per_km') or r.get('weight_kg_per_km') or 0)
		remarks = r.get('remarks') or r.get('note') or ''

		key = (vendor, part, spec, cores, float(csa))
		agg[key]['length'] += length
		agg[key]['weight'] += (weight_per_km * (length / 1000.0))
		if remarks:
			agg[key]['remarks'].add(str(remarks))
		agg[key]['qty_count'] += 1

	# Write rows
	for key, vals in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
		vendor, part, spec, cores, csa = key
		total_len = round(vals['length'], 2)
		total_wt = round(vals['weight'], 3)
		remarks = '; '.join(sorted(vals['remarks'])) if vals['remarks'] else ''
		ws.append([vendor, part, spec, cores, csa, total_len, total_wt, remarks])
		for cell in ws[ws.max_row]:
			cell.border = thin_border
			cell.alignment = Alignment(horizontal='center', vertical='center')

	# Column widths
	for col, w in zip(('A','B','C','D','E','F','G','H'), (18,18,30,8,12,12,14,30)):
		ws.column_dimensions[col].width = w

	# Save to bytes and file
	excel_bytes = BytesIO()
	wb.save(excel_bytes)
	excel_bytes.seek(0)

	timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
	filename = f"BOQ_{project}_{timestamp}.xlsx"
	filepath = f"/tmp/{filename}"
	with open(filepath, 'wb') as f:
		f.write(excel_bytes.getvalue())

	return {'downloadUrl': f'/cable/download-excel/{filename}', 'filename': filename}


@router.post('/export/sizing-report', response_model=dict)
def export_sizing_report(payload: dict):
	"""Generate a detailed sizing report XLSX. Payload: { rows: [...], columns: [...] , project: optional }
	Returns downloadUrl and filename.
	"""
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
		from datetime import datetime
	except ImportError:
		raise HTTPException(status_code=500, detail='openpyxl not installed')

	rows = payload.get('rows', [])
	columns = payload.get('columns', None)
	project = payload.get('project', 'PROJECT')
	if not rows:
		raise HTTPException(status_code=400, detail='No rows provided')

	wb = Workbook()
	if 'Sheet' in wb.sheetnames:
		wb.remove(wb['Sheet'])

	ws = wb.create_sheet('Sizing Report', 0)

	# Default columns if none provided
	default_cols = [
		'Cable No', 'From', 'To', 'Voltage', 'Load kW', 'Length m',
		'FLC A', 'Derated A', 'CSA mm2', 'Vdrop %', 'Start Vdrop %', 'Start Method',
		'SC OK', 'SC Required Area', 'Catalog Vendor', 'Catalog Part', 'Remarks'
	]

	headers = columns if columns and isinstance(columns, list) else default_cols
	ws.append(headers)

	header_fill = PatternFill(start_color='0B5394', end_color='0B5394', fill_type='solid')
	header_font = Font(bold=True, color='FFFFFF')
	thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal='center', vertical='center')

	for r in rows:
		result = r.get('result', {}) or {}
		cable_no = r.get('cable_number', '')
		from_eq = r.get('from_equipment', '')
		to_eq = r.get('to_equipment', '')
		voltage = r.get('voltage', '')
		load_kw = r.get('load_kw', '')
		length = r.get('length', 0)
		flc = result.get('flc', '')
		derated = result.get('derated_current', '')
		csa = result.get('selected_csa', '')
		vdrop = result.get('vdrop_percent', '')
		# find start vdrop in compliance
		start_item = None
		sc_item = None
		if result.get('compliance'):
			for ci in result['compliance']:
				if ci.get('type') in ('vdrop_start', 'vdrop_starting', 'vdrop_start'):
					start_item = ci
				if ci.get('type') == 'sc':
					sc_item = ci

		start_vdrop = start_item.get('value') if start_item else ''
		start_method = r.get('motor_start_method') or (start_item.get('msg') if start_item else '')
		sc_ok = result.get('sc_ok', '')
		sc_req = result.get('sc_required_area', '')
		vendor = r.get('catalog_vendor') or ''
		part = r.get('catalog_part_no') or ''
		remarks = r.get('remarks') or ''

		row_out = []
		for h in headers:
			key = h.lower()
			if h == 'Cable No': row_out.append(cable_no)
			elif h == 'From': row_out.append(from_eq)
			elif h == 'To': row_out.append(to_eq)
			elif h == 'Voltage': row_out.append(voltage)
			elif h == 'Load kW': row_out.append(load_kw)
			elif h == 'Length m': row_out.append(length)
			elif h in ('FLC A', 'FLC'): row_out.append(flc)
			elif h in ('Derated A', 'Derated'): row_out.append(derated)
			elif h in ('CSA mm2', 'CSA (mm²)'): row_out.append(csa)
			elif h in ('Vdrop %', 'Vdrop'): row_out.append(vdrop)
			elif h in ('Start Vdrop %', 'Start Vdrop'): row_out.append(start_vdrop)
			elif h == 'Start Method': row_out.append(start_method)
			elif h == 'SC OK': row_out.append('Yes' if sc_ok else 'No')
			elif h == 'SC Required Area': row_out.append(sc_req)
			elif h == 'Catalog Vendor': row_out.append(vendor)
			elif h == 'Catalog Part': row_out.append(part)
			elif h == 'Remarks': row_out.append(remarks)
			else:
				# try to fetch generic keys
				row_out.append(r.get(h) or result.get(h) or '')

		ws.append(row_out)
		for cell in ws[ws.max_row]:
			cell.border = thin_border
			cell.alignment = Alignment(horizontal='center', vertical='center')

	# autosize some columns
	for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
		ws.column_dimensions[col[0].column_letter].width = 18

	excel_bytes = BytesIO()
	wb.save(excel_bytes)
	excel_bytes.seek(0)

	timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
	filename = f"SizingReport_{project}_{timestamp}.xlsx"
	filepath = f"/tmp/{filename}"
	with open(filepath, 'wb') as f:
		f.write(excel_bytes.getvalue())

	return {'downloadUrl': f'/cable/download-excel/{filename}', 'filename': filename}


@router.get('/download-excel/{filename}')
def download_excel(filename: str):
	"""Download the generated Excel file."""
	filepath = f"/tmp/{filename}"
	if not os.path.exists(filepath):
		raise HTTPException(status_code=404, detail='File not found')
	
	return FileResponse(
		filepath,
		media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		filename=filename
	)



@router.post('/upload', response_model=dict)
def upload_excel(file: UploadFile = File(...)):
	"""Upload an Excel file and return detected headers, sample rows and a token to map later."""
	if not file.filename.lower().endswith(('.xls', '.xlsx')):
		raise HTTPException(status_code=400, detail='Only Excel files are supported')

	token = str(uuid.uuid4())
	save_path = f"/tmp/{token}.xlsx"
	with open(save_path, 'wb') as f:
		f.write(file.file.read())

	try:
		df = pd.read_excel(save_path, engine='openpyxl')
	except Exception as e:
		raise HTTPException(status_code=400, detail=f'Failed to read Excel: {e}')

	headers = [str(c) for c in df.columns.tolist()]
	sample = df.head(10).fillna('').to_dict(orient='records')

	return {'token': token, 'headers': headers, 'sample': sample}



@router.post('/catalog/upload', response_model=dict)
def catalog_upload(file: UploadFile = File(...)):
	"""Upload a vendor catalog Excel and return token + headers + sample"""
	if not file.filename.lower().endswith(('.xls', '.xlsx')):
		raise HTTPException(status_code=400, detail='Only Excel files are supported')

	token = str(uuid.uuid4())
	save_path = get_catalog_excel_path(token)
	with open(save_path, 'wb') as f:
		f.write(file.file.read())

	try:
		info = read_excel_headers_sample(save_path)
	except Exception as e:
		raise HTTPException(status_code=400, detail=f'Failed to read Excel: {e}')

	return {'token': token, 'headers': info['headers'], 'sample': info['sample']}



@router.post('/catalog/map', response_model=dict)
def catalog_map(payload: dict):
	"""Accepts { token: str, mapping: { canonicalField: columnName } } and stores normalized catalog on disk."""
	token = payload.get('token')
	mapping = payload.get('mapping') or {}
	if not token:
		raise HTTPException(status_code=400, detail='token required')

	excel_path = get_catalog_excel_path(token)
	if not os.path.exists(excel_path):
		raise HTTPException(status_code=404, detail='Uploaded catalog not found')

	try:
		df = pd.read_excel(excel_path, engine='openpyxl')
	except Exception as e:
		raise HTTPException(status_code=400, detail=f'Failed to read saved Excel: {e}')

	entries = normalize_catalog(df, mapping)

	# Save normalized to JSON for session use
	save_json = get_catalog_path(token)
	with open(save_json, 'w', encoding='utf-8') as f:
		json.dump(entries, f, indent=2)

	return {'token': token, 'count': len(entries), 'preview': entries[:10]}



@router.get('/catalog/list/{token}', response_model=list)
def catalog_list(token: str):
	"""Return normalized catalog entries for a given token."""
	data = load_catalog(token)
	if not data:
		raise HTTPException(status_code=404, detail='Catalog not found')
	return data



@router.post('/catalog/match', response_model=list)
def catalog_match(payload: dict):
	"""Match rows to catalog entries. Payload: { token, rows: [ { cable_number, voltage, conductor, derated_current, csa_preference(optional) } ], top_n: 3 }
	Returns list of suggestions per row.
	"""
	token = payload.get('token')
	rows = payload.get('rows', [])
	top_n = int(payload.get('top_n', 3))

	if not token:
		raise HTTPException(status_code=400, detail='token required')

	catalog = load_catalog(token)
	if not catalog:
		raise HTTPException(status_code=404, detail='Catalog not found')
	matches = []
	for row in rows:
		# Derated current may be provided; if not compute basic FLC
		derated = row.get('derated_current')
		if derated is None:
			# compute if load_kw or current provided
			if row.get('current'):
				i_base = float(row.get('current'))
			elif row.get('load_kw'):
				i_base = full_load_current_kw(float(row.get('load_kw')), float(row.get('voltage')), float(row.get('pf', 1.0)), float(row.get('eff', 1.0)))
			else:
				i_base = 0.0
			derated = derated_current(i_base, row.get('derating_factors', [1.0]))

		conductor = (row.get('conductor') or '').strip().lower()

		# filter candidates where csa_mm2 >= derated
		candidates = []
		for entry in catalog:
			c_csa = float(entry.get('csa_mm2') or 0)
			c_cond = (entry.get('conductor') or '').strip().lower()
			# simple match: conductor if provided, else accept
			if conductor and c_cond and conductor not in c_cond:
				continue
			if c_csa >= derated:
				candidates.append({'entry': entry, 'csa': c_csa})

		# sort by smallest csa that satisfies
		candidates_sorted = sorted(candidates, key=lambda x: x['csa'])[:top_n]
		suggestions = [c['entry'] for c in candidates_sorted]

		# Fallback: if no candidate meets derated, suggest top_n closest CSAs (either side)
		if not suggestions:
			all_candidates = []
			for entry in catalog:
				try:
					c_csa = float(entry.get('csa_mm2') or 0)
				except Exception:
					c_csa = 0.0
				all_candidates.append({'entry': entry, 'csa': c_csa, 'diff': abs(c_csa - derated)})
			suggestions = [c['entry'] for c in sorted(all_candidates, key=lambda x: x['diff'])[:top_n]]
		matches.append({'cable_number': row.get('cable_number'), 'suggestions': suggestions})

	return matches


@router.post('/map-upload', response_model=list)
def map_upload(payload: dict):
	"""Accepts { token: str, mapping: {targetField: columnName} } and returns array of BulkRow-like objects."""
	token = payload.get('token')
	mapping = payload.get('mapping') or {}
	if not token:
		raise HTTPException(status_code=400, detail='token required')

	path = f"/tmp/{token}.xlsx"
	if not os.path.exists(path):
		raise HTTPException(status_code=404, detail='Uploaded file not found')

	try:
		df = pd.read_excel(path, engine='openpyxl')
	except Exception as e:
		raise HTTPException(status_code=400, detail=f'Failed to read saved Excel: {e}')

	results = []
	for _, row in df.fillna('').iterrows():
		out = {
			'cable_number': str(row.get(mapping.get('cable_number', ''), '')),
			'from_equipment': str(row.get(mapping.get('from_equipment', ''), '')),
			'to_equipment': str(row.get(mapping.get('to_equipment', ''), '')),
			'load_kw': float(row.get(mapping.get('load_kw', ''), 0) or 0),
			'load_kva': float(row.get(mapping.get('load_kva', ''), 0) or 0),
			'current': float(row.get(mapping.get('current', ''), 0) or 0),
			'voltage': float(row.get(mapping.get('voltage', ''), 415) or 415),
			'pf': float(row.get(mapping.get('pf', ''), 1.0) or 1.0),
			'eff': float(row.get(mapping.get('eff', ''), 1.0) or 1.0),
			'length': float(row.get(mapping.get('length', ''), 0) or 0),
			'mv_per_a_m': float(row.get(mapping.get('mv_per_a_m', ''), 0.44) or 0.44),
			'derating1': float(row.get(mapping.get('derating1', ''), 1) or 1),
			'derating2': float(row.get(mapping.get('derating2', ''), 1) or 1),
			'sc_current': float(row.get(mapping.get('sc_current', ''), 0) or 0),
			'sc_time': float(row.get(mapping.get('sc_time', ''), 1) or 1),
			'k_const': float(row.get(mapping.get('k_const', ''), 115) or 115),
		}
		results.append(out)

	return results

